#!/usr/bin/env python
# coding: utf-8

# In[1]:


#https://openpyxl.readthedocs.io/en/stable/
from openpyxl import load_workbook


# In[2]:


#This program will open excel and will compare one column from all of the rows between two sheets (A and B)
#and create two new sheets.
#Each sheet will list every row from A with the rows from B under it that matched it


# In[3]:


#Read in config variables
import configparser

config = configparser.ConfigParser()
config.read_file(open('config.ini'))
excel_section = 'Excel Info'

Excel_Filename = config.get(excel_section, 'name')
Name_Of_Sheet_A = config.get(excel_section, 'Name_Of_Sheet_A')
Name_Of_Sheet_B = config.get(excel_section, 'Name_Of_Sheet_B')

Col_To_Join_On_A = int(config.get(excel_section, 'Col_To_Join_On_A'))
Col_To_Join_On_B = int(config.get(excel_section, 'Col_To_Join_On_B'))

Include_Not_Found = config.getboolean(excel_section, 'Include_Not_Found')

#Derived config vars. Don't touch
Name_Of_Sorted_Sheet_A = Name_Of_Sheet_A + " compared with " + Name_Of_Sheet_B
Name_Of_Sorted_Sheet_B = Name_Of_Sheet_B + " compared with " + Name_Of_Sheet_A


# In[5]:


#Opening the files and loading the variables
wb = load_workbook(Excel_Filename)
sheet_A = wb[Name_Of_Sheet_A]
sheet_B = wb[Name_Of_Sheet_B]

#delete previous sheets if exist:
try:
    wb.remove(wb[Name_Of_Sheet_A + " Sorted by Matcher"])
    wb.remove(wb[Name_Of_Sheet_B + " Sorted by Matcher"])
except:
    print("No previous sorted sheets!")
sheet_A_sorted = wb.create_sheet(Name_Of_Sheet_A + " Sorted by Matcher")
sheet_B_sorted = wb.create_sheet(Name_Of_Sheet_B + " Sorted by Matcher")


# In[6]:


#Turn the WS object into list form
def worksheet_to_list(ws):
    ws_values = ws.values
    return [list(row) for row in ws_values]

sheet_A_list_form = worksheet_to_list(sheet_A)
sheet_B_list_form = worksheet_to_list(sheet_B)


# In[7]:


#Isolate the columns that we're supposed to be matching on

#the_col_to_be_compared_in_sheet_b
col_B = [row[Col_To_Join_On_B] for row in sheet_B_list_form]
#the_col_to_be_compared_in_sheet_a 
col_A = [row[Col_To_Join_On_A] for row in sheet_A_list_form]


# In[8]:


#We're going to loop through every single item in col_A and see if there's a match in colB
#We will get the indicies (aka row number) for all rows in sheet B that matched
#We will copy/extract those rows from sheet_B

#and write them to a new sheet in form:
# Row from Sheet A:
# X X X X X
# Matching Rows from Sheet B:
# Y Y Y Y Y
# Y Y Y Y Y
# Y Y Y Y Y


# In[9]:


#This used to be free floating code. Turned it into a func to call it multiple times. So much
#easier than changing all the A's into B's and vice versa by hand.
#The tradeoff was that to do minimal changes, there are a lot of variable names that were written from
#A compared to B perspective. The local scope vars will share the same name as the vars one scope above it.
#Python should be able to resolve this just fine.

def the_engine(col_A, col_B, sheet_A_list_form, sheet_B_list_form, ws_to_write_to):
    #Looping through every single thing in A
    for index, item in enumerate(col_A):
        #Get the index of all rows in B that contain a match to our current item
        matches = [ i for i in range(len(col_B)) if col_B[i] == item]

        #Get the row from A that we're going to put in
        a_row = sheet_A_list_form[index]

        #print the relevant row from A first. BUT!....(see next line)
        #Do we include Sheet A rows without matches?
        if Include_Not_Found:
            #If include_not_found is true, always write the row from sheet A
            ws_to_write_to.append(a_row)
        else:
            #Otherwise check if matches has anything.
            #Only write row A if there are matches.
            #Otherwise go to next loop.
            if len(matches) > 0:
                ws_to_write_to.append(a_row)
            else:
                continue

        for b_index in matches:
            #I'm just splitting this up because it feels too verbose put together...
            b_row = sheet_B_list_form[b_index]
            ws_to_write_to.append(b_row)

        #Add a space at the end to separate each batch of 
        #ZZZZZZ
        #YYYYYY
        #YYYYYY
        ws_to_write_to.append([])    

the_engine(col_A, col_B, sheet_A_list_form,sheet_B_list_form, sheet_A_sorted)
the_engine(col_B, col_A, sheet_B_list_form,sheet_A_list_form, sheet_B_sorted)


# In[10]:


#The below uses a row counter and manually inserting data. The code is incomplete
#The above will attempt to use append. Decided to try append approach instead of manually putting data in cells.
    
# #A counter for which row we're on in the compared_sheet. Starts at 1
# #because excel starts at one.
# row_counter = 1 

# #Looping through every single thing in A
# for index, item in enumerate(col_A):
#     #Get the index of all rows in B that contain a match to our current item
#     matches = [ i for i in range(len(col_B)) if col_B[i] == item]
#     print (item)
#     #print the relevant row from A first
#     sheet_A_sorted.cell(row = row_counter, column=1).value = sheet_A_list_form[index]
    
#     #Add a space between row from Sheet A and the found rows from sheetB
#     row_counter += 1
    
#     #Add 1 to the row counter in the end to add a space between each result
#     row_counter += 1


# In[11]:


#This saves all of our changes above into the file
wb.save(Excel_Filename)


#https://openpyxl.readthedocs.io/en/stable/
from openpyxl import load_workbook

#This program will open excel and will compare one column from all of the rows between two sheets (A and B) and produce 3 new sheets 
#Based on the column to join on
#1 - The rows in Sheet A but not Sheet B
#2 - The rows in Sheet B but not Sheet A
#3 - The 

#This is basically a double outer join for #1 and #2, and an inner join but a bit more 

#Story: Person A says "Hey I sent you 3 checks. Here's an excel document of their date, amount,
# check number and check description".
#Person B says "Hey, I only see two checks. I'm going to run this program with Person A's checks and my list of received checks. 
# Whichever checks in person 
# matching pair means I did not get it!"

#Ultimately, this is just a subtraction function of Person A's list minus Person B's list/

#There are two main sheets: Dictionary and Searchtionary

#Config Vars here:
Excel_Filename = "test.xlsx"
Name_Of_Sheet_A = ""
Name_Of_Sheet_B = ""
Col_To_Join_On = ""
    #


#Opening the files and loading the variables
work_book = load_workbook(Excel_Filename)
sheet_A = 



sheet_names = work_book.sheetnames
dictionary_sheet = work_book[Sheet_Name_For_Dictionary_Sheet]




print (a1 , b1)
print ("is a1 == b1? ", a1 == b1)
#Don't forget to save the workbook!
work_book.save('test.xlsx')

